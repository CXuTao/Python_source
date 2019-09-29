# xlsx  xls
#先要安装 openpyxl -> xlsx  pip install openpyxl 安装时原来的下载网址出现问题
#换源下载pip install --index https://mirrors.ustc.edu.cn/pypi/web/simple/ openpyxl

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    #print(file.get_sheet_names())该方法python3.7已经弃用，应该使用以下方法
    sheets = file.sheetnames

    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    #最大行数
    #print(sheet.max_row)
    #最大列数
    #print(sheet.max_column)
    #表名
    #print(sheet.title)

    #读取数据
    for lineNum in range(1, sheet.max_row + 1):
        #print(lineNum)
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            #拿数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            """if value != None:
                lineList.append(value)"""
            lineList.append(value)
        print(lineList)







path = r"E:\python\office文件操作\4.excel操作\01.xlsx"
readXlsxFile(path)