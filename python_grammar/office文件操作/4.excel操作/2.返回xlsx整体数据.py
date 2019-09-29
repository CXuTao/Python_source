# xlsx  xls
#先要安装 openpyxl -> xlsx  pip install openpyxl 安装时原来的下载网址出现问题
#换源下载

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    dic = {}
    file = load_workbook(filename=path)
    sheets = file.sheetnames

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)

        #将一张表的数据存到字典
        dic[sheetName] = sheetInfo
    return dic


path = r"E:\python\office文件操作\4.excel操作\01.xlsx"
dic = readXlsxFile(path)
print(len(dic))